import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import pdfplumber
import re
import io
import math

st.set_page_config(page_title="Asignador CE | INTERLOG", page_icon="📜", layout="wide")

st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Roboto:wght@300;400;500;700&display=swap');
    html, body, [class*="css"] { font-family: 'Roboto', sans-serif; background-color: #0e1117; color: #fafafa; }
    .block-container { padding-top: 3rem; }
    h1 { color: #00b4d8 !important; font-size: 1.8rem !important; }
    .subtitulo { font-size: 0.95rem; color: #8899aa; margin-bottom: 1.5rem; }
    .seccion { font-size: 1rem; font-weight: 600; color: #00b4d8; text-transform: uppercase; letter-spacing: 1px; margin-bottom: 1rem; border-bottom: 1px solid #1e3a4a; padding-bottom: 6px; }
    .badge-ok { background: #0d2e1a; color: #3dd68c; padding: 5px 12px; border-radius: 20px; font-size: 12px; font-weight: 500; border: 1px solid #3dd68c; display: inline-block; margin: 3px 2px; }
    .badge-err { background: #2e0d0d; color: #ff6b6b; padding: 5px 12px; border-radius: 20px; font-size: 12px; font-weight: 500; border: 1px solid #ff6b6b; display: inline-block; margin: 3px 2px; }
    .badge-warn { background: #2e2200; color: #f0c040; padding: 5px 12px; border-radius: 20px; font-size: 12px; font-weight: 500; border: 1px solid #f0c040; display: inline-block; margin: 3px 2px; }
    .match-box { background: #161b22; border: 1px solid #1e3a4a; border-radius: 8px; padding: 14px; margin: 6px 0; font-size: 0.88rem; }
    .match-ok { border-color: #3dd68c; }
    .match-warn { border-color: #f0c040; }
    .match-err { border-color: #ff6b6b; }
    .stButton > button { background: linear-gradient(135deg, #00b4d8, #0077b6) !important; color: white !important; font-weight: 600 !important; border: none !important; border-radius: 8px !important; }
    #MainMenu {visibility: hidden;} header {visibility: hidden;} footer {visibility: hidden;}
</style>
""", unsafe_allow_html=True)

def extraer_nro_ce(filename):
    m = re.search(r'(CE-\d{4}-\d+-APN-DIMI)', filename)
    return m.group(1) + '#MEC' if m else None

def extraer_nro_re(filename):
    m = re.search(r'(RE-\d{4}-\d+-APN-DGDA)', filename)
    return m.group(1) + '#MEC' if m else None

def extraer_texto_pdf(file):
    texto = ''
    with pdfplumber.open(file) as pdf:
        for page in pdf.pages:
            texto += (page.extract_text() or '') + ' '
    return texto

def extraer_re_de_ce(texto):
    m = re.search(r'RE-\d{4}-\d+-APN-[\s\n]*DGDA#MEC', texto)
    return m.group(0).replace('\n', '').replace(' ', '') if m else None

def extraer_datos_re(texto):
    fob_m = re.search(r'Valor FOB TOTAL[^\d]*([\d,\.]+)', texto)
    fob_str = fob_m.group(1) if fob_m else '0'
    fob = float(fob_str.replace('.', '').replace(',', '.'))
    codigos = re.findall(r'Código de parte[^:]*:\s*([A-Z0-9]+)', texto)
    codigos = [c for c in codigos if c.upper() not in ('NOPOSES', 'NO') and len(c) > 2]
    return fob, codigos

def safe_float(v):
    try:
        f = float(str(v).replace(',', '.'))
        return 0.0 if math.isnan(f) else f
    except:
        return 0.0

def exportar_excel(df):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'general'
    FILL_AM = PatternFill('solid', fgColor='FFFF00')
    FH = Font(name='Arial', size=11, bold=True)
    FD = Font(name='Calibri', size=11)
    SIN_COLOR = {'ID','InscRUMP','ActiServ','NroInsc','RazonSocial','CUIT','ImpDirecta','CondMerca',
                 'SimiSira','ProyectoMinero','Radicacion','ClasificacionDeArticulo','TipoDeFactura',
                 'Observaciones','ITEM_DESPACHO','NRO_CE'}
    for ci, col in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.font = FH; cell.alignment = Alignment(horizontal='left', vertical='center')
        if col not in SIN_COLOR: cell.fill = FILL_AM
    for ri, row in enumerate(df.itertuples(index=False), 2):
        for ci, val in enumerate(row, 1):
            v = None if (isinstance(val, float) and math.isnan(val)) else val
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.font = FD; cell.alignment = Alignment(vertical='center')
    for col in ws.columns:
        ml = max((len(str(c.value)) if c.value is not None else 0) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max(ml + 2, 10), 45)
    ws.freeze_panes = 'A2'
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.getvalue()

def asignar_ce(df, ce_info):
    """
    Asigna CE a filas del unificado.
    Lógica: para cada CE busca filas con los códigos de parte del RE
    y verifica que la suma de FOB coincida con el FOB del RE.
    """
    df = df.copy()
    if 'NRO_CE' not in df.columns:
        idx = df.columns.tolist().index('ITEM_DESPACHO') + 1
        df.insert(idx, 'NRO_CE', '')
    
    # Solo filas que aplican CM (sin SIN CM, sin CE ya asignado)
    mask_aplica = df['Observaciones'].isna() | (df['Observaciones'].str.strip() == '')
    
    resultados = []
    alertas_duplicados = []
    
    # Detectar RE con mismos codigos Y mismo FOB (ambigüedad)
    firma_re = {}  # firma -> lista de CEs
    for nro_ce, info in ce_info.items():
        firma = (frozenset(info['codigos']), round(info['fob'], 2))
        firma_re.setdefault(firma, []).append(nro_ce)
    
    for firma, ces_lista in firma_re.items():
        if len(ces_lista) > 1:
            alertas_duplicados.append({
                'ces': ces_lista,
                'codigos': list(firma[0]),
                'fob': firma[1]
            })
    
    for nro_ce, info in ce_info.items():
        codigos_ce = info['codigos']
        fob_ce = info['fob']
        
        # Filas disponibles (aplican CM y aún sin CE asignado)
        mask_disp = mask_aplica & (df['NRO_CE'] == '')
        candidatos = df[mask_disp & df['CodigoParte'].isin(codigos_ce)].copy()
        
        if candidatos.empty:
            resultados.append({'ce': nro_ce, 're': info['re'], 'estado': 'sin_match',
                               'fob_ce': fob_ce, 'fob_calc': 0, 'n_items': 0})
            continue
        
        # Verificar FOB
        fob_calc = round(sum(safe_float(v) for v in candidatos['ValorTotalItem']), 2)
        diff = abs(fob_calc - fob_ce)
        estado = 'ok' if diff <= 1 else 'warn'
        
        df.loc[candidatos.index, 'NRO_CE'] = nro_ce
        
        resultados.append({'ce': nro_ce, 're': info['re'], 'estado': estado,
                           'fob_ce': fob_ce, 'fob_calc': fob_calc, 'n_items': len(candidatos)})
    
    return df, resultados, alertas_duplicados

# ── UI ────────────────────────────────────────────────────────────────────────
st.title("📜 Asignador CE")
st.markdown('<p class="subtitulo">INTERLOG Comercio Exterior — Asignación de Certificados Mineros al Template Unificado</p>', unsafe_allow_html=True)
st.markdown("---")

col1, col2 = st.columns([1, 1])
with col1:
    st.markdown('<p class="seccion">📊 Template Unificado</p>', unsafe_allow_html=True)
    f_unificado = st.file_uploader("Template Unificado (.xlsx)", type=["xlsx"])
with col2:
    st.markdown('<p class="seccion">📄 PDFs de CE y RE</p>', unsafe_allow_html=True)
    f_pdfs = st.file_uploader("PDFs de CE y RE (múltiples)", type=["pdf"], accept_multiple_files=True)

st.markdown("---")

if not (f_unificado and f_pdfs):
    st.info("📝 Subí el Template Unificado y los PDFs de CE y RE para continuar.")

if st.button("🔍 ANALIZAR Y ASIGNAR CE", disabled=not (f_unificado and f_pdfs), use_container_width=True):
    with st.spinner("Procesando..."):
        df = pd.read_excel(f_unificado, dtype=str)
        df.columns = df.columns.str.strip()
        
        ces = {}; res = {}
        for f in f_pdfs:
            fname = f.name
            texto = extraer_texto_pdf(f)
            if 'CE-' in fname and 'DIMI' in fname:
                nro_ce = extraer_nro_ce(fname)
                nro_re = extraer_re_de_ce(texto)
                if nro_ce and nro_re:
                    ces[nro_ce] = nro_re
            elif 'RE-' in fname and 'DGDA' in fname:
                nro_re = extraer_nro_re(fname)
                fob, codigos = extraer_datos_re(texto)
                if nro_re:
                    res[nro_re] = {'fob': fob, 'codigos': codigos}
        
        # Construir ce_info solo con CEs que tienen su RE
        ce_info = {}
        for nro_ce, nro_re in ces.items():
            if nro_re in res:
                ce_info[nro_ce] = {'re': nro_re, 'fob': res[nro_re]['fob'], 'codigos': res[nro_re]['codigos']}
        
        df_resultado, resultados, alertas_dup = asignar_ce(df, ce_info)
        
        st.session_state.update({
            'df_resultado': df_resultado,
            'resultados': resultados,
            'alertas_dup': alertas_dup,
            'fname_original': f_unificado.name,
            'procesado': True
        })

if st.session_state.get('procesado'):
    resultados = st.session_state['resultados']
    df = st.session_state['df_resultado']
    alertas_dup = st.session_state['alertas_dup']
    fname = st.session_state['fname_original']
    
    # Alertas de duplicados
    if alertas_dup:
        st.markdown("### ⚠️ Alertas — CEs con mismos códigos y FOB")
        for a in alertas_dup:
            ces_str = ' / '.join(a['ces'])
            st.markdown(f'<span class="badge-warn">⚠️ CEs ambiguos: {ces_str} | FOB: {a["fob"]} | Códigos: {len(a["codigos"])}</span>', unsafe_allow_html=True)
    
    st.markdown("### 📊 Resultado de la asignación")
    for r in resultados:
        if r['estado'] == 'ok':
            css, badge = 'match-ok', '<span class="badge-ok">✅ OK</span>'
        elif r['estado'] == 'warn':
            css, badge = 'match-warn', '<span class="badge-warn">⚠️ Diferencia FOB</span>'
        else:
            css, badge = 'match-err', '<span class="badge-err">❌ Sin match en unificado</span>'
        
        st.markdown(f"""
        <div class="match-box {css}">
            {badge} <b>{r['ce']}</b><br>
            RE: {r['re']}<br>
            Ítems asignados: {r['n_items']} | FOB RE: USD {r['fob_ce']} | FOB calculado: USD {r['fob_calc']}
        </div>
        """, unsafe_allow_html=True)
    
    # Items sin CE
    sin_ce = df[(df['NRO_CE'] == '') & (df['Observaciones'].isna() | (df['Observaciones'].str.strip() == ''))]
    if not sin_ce.empty:
        st.markdown(f'<br><span class="badge-warn">⚠️ {len(sin_ce)} ítems sin CE asignado</span>', unsafe_allow_html=True)
        st.dataframe(sin_ce[['NumeroDeFactura','CodigoParte','ValorTotalItem','ITEM_DESPACHO']].reset_index(drop=True))
    
    total_asignados = (df['NRO_CE'] != '').sum()
    total_sin_cm = (df['Observaciones'].str.strip() == 'SIN CM').sum() if 'Observaciones' in df.columns else 0
    st.markdown(f"**Total: {total_asignados} ítems con CE | {len(sin_ce)} sin CE | {total_sin_cm} SIN CM (esperado)**")
    
    st.markdown("---")
    st.download_button(
        label="📥 DESCARGAR UNIFICADO CON CE",
        data=exportar_excel(df),
        file_name=fname.replace('.xlsx', '_CON_CE.xlsx'),
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True
    )
