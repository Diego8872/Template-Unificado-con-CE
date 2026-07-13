import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import pdfplumber
import re
import io
import math
from itertools import combinations

st.set_page_config(page_title="Template Unificado con CE | INTERLOG", page_icon="📜", layout="wide")

st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Roboto:wght@300;400;500;700&display=swap');
    html, body, [class*="css"] { font-family: 'Roboto', sans-serif; background-color: #0e1117; color: #fafafa; }
    .block-container { padding-top: 3rem; }
    h1 { color: #00b4d8 !important; font-size: 1.8rem !important; }
    h2, h3 { color: #00b4d8 !important; }
    .subtitulo { font-size: 0.95rem; color: #8899aa; margin-bottom: 1.5rem; }
    .seccion { font-size: 1rem; font-weight: 600; color: #00b4d8; text-transform: uppercase; letter-spacing: 1px; margin-bottom: 1rem; border-bottom: 1px solid #1e3a4a; padding-bottom: 6px; }
    .badge-ok { background: #0d2e1a; color: #3dd68c; padding: 5px 12px; border-radius: 20px; font-size: 12px; font-weight: 500; border: 1px solid #3dd68c; display: inline-block; margin: 3px 2px; }
    .badge-err { background: #2e0d0d; color: #ff6b6b; padding: 5px 12px; border-radius: 20px; font-size: 12px; font-weight: 500; border: 1px solid #ff6b6b; display: inline-block; margin: 3px 2px; }
    .badge-warn { background: #2e2200; color: #f0c040; padding: 5px 12px; border-radius: 20px; font-size: 12px; font-weight: 500; border: 1px solid #f0c040; display: inline-block; margin: 3px 2px; }
    .match-box { background: #161b22; border: 1px solid #1e3a4a; border-radius: 8px; padding: 14px; margin: 6px 0; font-size: 0.92rem; color: #e0e0e0; }
    .match-ok { border-color: #3dd68c; }
    .match-warn { border-color: #f0c040; }
    .match-err { border-color: #ff6b6b; }
    .match-box b { color: #ffffff; }
    .match-box span { color: #aaaaaa; }
    .stButton > button { background: linear-gradient(135deg, #00b4d8, #0077b6) !important; color: white !important; font-weight: 600 !important; border: none !important; border-radius: 8px !important; }
    #MainMenu {visibility: hidden;} header {visibility: hidden;} footer {visibility: hidden;}
</style>
""", unsafe_allow_html=True)

def norm_codigo(cod):
    return str(cod).strip().lstrip('0') if cod else ''

def extraer_nro_ce_desde_pdf(file):
    try:
        with pdfplumber.open(file) as pdf:
            for page in pdf.pages:
                for annot in (page.annots or []):
                    data = annot.get('data', {})
                    titulo = data.get('T', b'')
                    if isinstance(titulo, bytes):
                        titulo = titulo.decode('latin-1', errors='ignore')
                    if 'numero_documento' in titulo.lower():
                        valor = data.get('V', b'')
                        if isinstance(valor, bytes):
                            valor = valor.decode('latin-1', errors='ignore')
                        if valor and valor.startswith('CE-'):
                            return valor.strip()
    except Exception:
        pass
    return None

def extraer_nro_ce_desde_nombre(filename):
    m = re.search(r'(CE-\d{4}-\d+-APN-\w+?)(?:_|\b)', filename)
    if not m:
        return None
    nro = m.group(1)
    if not nro.endswith('#MEC'):
        nro = nro + '#MEC'
    return nro

def extraer_nro_re(filename):
    m = re.search(r'(RE-\d{4}-\d+-APN-DGDA)', filename)
    return m.group(1) + '#MEC' if m else None

def extraer_texto_pdf(file):
    texto = ''
    with pdfplumber.open(file) as pdf:
        for page in pdf.pages:
            texto += (page.extract_text() or '') + ' '
    return texto

def extraer_texto_y_nro_ce(file):
    texto = ''
    nro_ce = None
    try:
        with pdfplumber.open(file) as pdf:
            for page in pdf.pages:
                texto += (page.extract_text() or '') + ' '
                if nro_ce is None:
                    for annot in (page.annots or []):
                        data = annot.get('data', {})
                        titulo = data.get('T', b'')
                        if isinstance(titulo, bytes):
                            titulo = titulo.decode('latin-1', errors='ignore')
                        if 'numero_documento' in titulo.lower():
                            valor = data.get('V', b'')
                            if isinstance(valor, bytes):
                                valor = valor.decode('latin-1', errors='ignore')
                            if valor and valor.startswith('CE-'):
                                nro_ce = valor.strip()
    except Exception:
        pass
    return texto, nro_ce

def extraer_re_de_ce(texto):
    # [^\w\s]* acepta cualquier carácter especial entre # y MEC (ej: punto medio ·)
    m = re.search(r'(RE-\d{4}-\d+-APN-[\s\n]*DGDA)#[^\w\s]*MEC', texto)
    if m:
        return m.group(1).replace('\n', '').replace(' ', '') + '#MEC'
    return None

def extraer_datos_re(texto):
    fob_m = re.search(r'Valor FOB TOTAL[^\d]*([\d,\.]+)', texto)
    fob_str = fob_m.group(1) if fob_m else '0'
    fob = float(fob_str.replace('.', '').replace(',', '.'))

    factura_m = re.search(r'Número de Factura:\s*([A-Z0-9]+)', texto)
    factura = factura_m.group(1).strip() if factura_m else None

    renglones_raw = re.findall(
        r'Cantidad:\s*([\d,\.]+).*?'
        r'Valor total de los art[íi]culo/item \(FOB en d[óo]lares estadounidenses\):\s*([\d,\.]+).*?'
        r'Código de parte[^:]*:\s*([A-Z0-9]+)',
        texto, flags=re.DOTALL
    )

    renglones = []
    codigos = []
    cantidad_por_codigo = {}
    for cant_str, fob_renglon_str, cod in renglones_raw:
        cod = norm_codigo(cod)
        if not cod or cod.upper() in ('NOPOSES', 'NO') or len(cod) <= 1:
            continue
        cant = safe_float(cant_str)
        fob_renglon = safe_float(fob_renglon_str)
        renglones.append({'codigo': cod, 'cantidad': cant, 'fob': fob_renglon})
        codigos.append(cod)
        cantidad_por_codigo[cod] = cantidad_por_codigo.get(cod, 0) + cant

    return fob, codigos, factura, cantidad_por_codigo, renglones

def safe_float(v):
    try:
        f = float(str(v).replace(',', '.'))
        return 0.0 if math.isnan(f) else f
    except:
        return 0.0

def exportar_excel(df):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'general'
    FILL_AM = PatternFill('solid', fgColor='FFFF00')
    FH = Font(name='Arial', size=11, bold=True)
    FD = Font(name='Calibri', size=11)
    SIN_COLOR = {'ID','InscRUMP','ActiServ','NroInsc','RazonSocial','CUIT','ImpDirecta','CondMerca',
                 'SimiSira','ProyectoMinero','Radicacion','ClasificacionDeArticulo','TipoDeFactura',
                 'Observaciones','ITEM_DESPACHO','ITEM','D:CERTSM','V:AUTOLIQCONTRIMP'}
    for ci, col in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.font = FH; cell.alignment = Alignment(horizontal='left', vertical='center')
        if col not in SIN_COLOR: cell.fill = FILL_AM
    for ri, row in enumerate(df.itertuples(index=False), 2):
        for ci, val in enumerate(row, 1):
            v = None if (isinstance(val, float) and math.isnan(val)) else val
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.font = FD; cell.alignment = Alignment(vertical='center')
    for col in ws.columns:
        ml = max((len(str(c.value)) if c.value is not None else 0) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max(ml + 2, 10), 45)
    ws.freeze_panes = 'A2'
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.getvalue()

def preasignar_codigos_compartidos(df, ce_info):
    """
    Resuelve códigos de parte que compiten entre varios CEs, a nivel de
    renglón individual (cantidad + FOB del renglón).

    FIX: al buscar candidatas para cada renglón, filtra por la factura del
    CE específico — no la unión de facturas de todos los competidores.
    Esto evita que un CE de factura X le robe una línea de factura Y a otro CE.
    """
    avisos = []

    # 1) Detectar qué códigos aparecen en más de un CE
    codigo_a_ces = {}
    for nro_ce, info in ce_info.items():
        for cod in set(info['codigos']):
            codigo_a_ces.setdefault(cod, []).append(nro_ce)
    codigos_compartidos = {c for c, ces_list in codigo_a_ces.items() if len(ces_list) > 1}

    if not codigos_compartidos:
        return df, avisos

    mask_aplica = df['Observaciones'].isna() | (df['Observaciones'].str.strip() == '')

    for cod in codigos_compartidos:
        ces_que_compiten = codigo_a_ces[cod]

        # Pool de líneas del Excel disponibles para este código (sin filtro de factura,
        # porque cada CE filtrará por la suya propia en el loop interno)
        mask_disp = mask_aplica & (df['D:CERTSM'] == '') & (df['CodigoParte'] == cod)
        candidatas = df[mask_disp].copy()
        if candidatas.empty:
            continue

        candidatas['_cant'] = candidatas['Cantidad'].apply(safe_float) if 'Cantidad' in candidatas.columns else 1.0
        candidatas['_fob'] = candidatas['ValorTotalItem'].apply(safe_float)

        pool = {i: (candidatas.at[i, '_cant'], candidatas.at[i, '_fob']) for i in candidatas.index}

        renglones_por_ce = {}
        for nro_ce in ces_que_compiten:
            renglones_por_ce[nro_ce] = [
                r for r in ce_info[nro_ce].get('renglones', []) if r['codigo'] == cod
            ]

        # FIX 1: filtrar por factura del CE específico
        # FIX 2: filtrar por ValorFOBTotal de la línea ≈ FOB del CE (discrimina líneas idénticas)
        for nro_ce, renglones in renglones_por_ce.items():
            factura_ce = ce_info[nro_ce].get('factura')
            fob_ce = ce_info[nro_ce]['fob']
            for rg in renglones:
                candidatos_idx = [
                    i for i, (c, f) in pool.items()
                    if abs(c - rg['cantidad']) < 0.001
                    and abs(f - rg['fob']) <= 1
                    and (not factura_ce or df.at[i, 'NumeroDeFactura'] == factura_ce)
                    and (
                        'ValorFOBTotal' not in df.columns
                        or pd.isna(df.at[i, 'ValorFOBTotal'])
                        or abs(safe_float(df.at[i, 'ValorFOBTotal']) - fob_ce) <= 1
                    )
                ]
                if candidatos_idx:
                    idx = candidatos_idx[0]
                    df.loc[idx, 'D:CERTSM'] = nro_ce
                    del pool[idx]
                else:
                    avisos.append(
                        f"⚠️ Código {cod}: no hay línea disponible en el Excel para "
                        f"{nro_ce} (cantidad {rg['cantidad']}, FOB {rg['fob']}). "
                        f"Verificar que el Template tenga esa línea."
                    )

    # Validación: si la pre-asignación dejó a un CE sin poder cerrar su FOB, revertir
    for nro_ce, info in ce_info.items():
        fob_ce = info['fob']
        factura = info.get('factura')
        idx_preasignados = df[df['D:CERTSM'] == nro_ce].index.tolist()
        fob_preasignado = round(sum(safe_float(df.at[i, 'ValorTotalItem']) for i in idx_preasignados), 2)
        if abs(fob_preasignado - fob_ce) <= 1:
            continue
        mask_disp = mask_aplica & (df['D:CERTSM'] == '') & (df['CodigoParte'].isin(info['codigos']))
        if factura:
            mask_disp = mask_disp & (df['NumeroDeFactura'] == factura)
        fob_disponible = round(sum(safe_float(v) for v in df[mask_disp]['ValorTotalItem']), 2)
        fob_total_posible = round(fob_preasignado + fob_disponible, 2)
        if abs(fob_total_posible - fob_ce) > 1:
            df.loc[idx_preasignados, 'D:CERTSM'] = ''
            avisos.append(
                f"ℹ️ Pre-asignación revertida para {nro_ce} "
                f"(FOB posible {fob_total_posible} ≠ FOB declarado {fob_ce}). "
                f"Se resolverá por matching global de FOB."
            )

    return df, avisos


def encontrar_subconjunto_fob(candidatos_df, fob_objetivo, tolerancia=1.0, timeout_seg=5):
    import time
    fobs = [safe_float(v) for v in candidatos_df['ValorTotalItem']]
    indices = list(candidatos_df.index)

    suma_total = round(sum(fobs), 2)
    if abs(suma_total - fob_objetivo) <= tolerancia:
        return indices, suma_total

    deadline = time.time() + timeout_seg
    if len(indices) <= 40:
        for size in range(len(indices)-1, 0, -1):
            if time.time() > deadline:
                break
            for combo_idx in combinations(range(len(indices)), size):
                if time.time() > deadline:
                    break
                suma = round(sum(fobs[i] for i in combo_idx), 2)
                if abs(suma - fob_objetivo) <= tolerancia:
                    return [indices[i] for i in combo_idx], suma

    return indices, suma_total


def asignar_ce(df, ce_info):
    df = df.copy()

    col_item = 'ITEM_DESPACHO' if 'ITEM_DESPACHO' in df.columns else (
        'ITEM' if 'ITEM' in df.columns else df.columns[-1]
    )
    if col_item != 'ITEM':
        df = df.rename(columns={col_item: 'ITEM'})
        col_item = 'ITEM'

    if 'D:CERTSM' not in df.columns:
        idx = df.columns.tolist().index(col_item) + 1
        df.insert(idx, 'D:CERTSM', '')

    mask_aplica = df['Observaciones'].isna() | (df['Observaciones'].str.strip() == '')

    firma_re = {}
    for nro_ce, info in ce_info.items():
        firma = (frozenset(info['codigos']), round(info['fob'], 2))
        firma_re.setdefault(firma, []).append(nro_ce)
    alertas_dup = [{'ces': v, 'fob': k[1]} for k, v in firma_re.items() if len(v) > 1]

    # PASO 1: pre-asignar códigos compartidos entre CEs (con fix de factura por CE)
    df, avisos_reparto = preasignar_codigos_compartidos(df, ce_info)

    resultados = []
    for nro_ce, info in ce_info.items():
        fob_ce = info['fob']

        ya_asignadas_idx = list(df[df['D:CERTSM'] == nro_ce].index)
        fob_acumulado = sum(safe_float(df.at[i, 'ValorTotalItem']) for i in ya_asignadas_idx)

        mask_disp = mask_aplica & (df['D:CERTSM'] == '') & (df['CodigoParte'].isin(info['codigos']))
        if info.get('factura'):
            mask_disp = mask_disp & (df['NumeroDeFactura'] == info['factura'])
        pool = df[mask_disp].copy()
        pool['_cant'] = pool['Cantidad'].apply(safe_float) if 'Cantidad' in pool.columns else 1.0
        pool['_fob'] = pool['ValorTotalItem'].apply(safe_float)

        # PASO 2: matching renglón-a-línea para códigos no compartidos
        renglones_pendientes = []
        for rg in info.get('renglones', []):
            ya = (
                (df['D:CERTSM'] == nro_ce) &
                (df['CodigoParte'] == rg['codigo']) &
                (df['Cantidad'].apply(safe_float).sub(rg['cantidad']).abs() < 0.001) &
                (df['ValorTotalItem'].apply(safe_float).sub(rg['fob']).abs() <= 1)
            ).any()
            if ya:
                continue

            mask_cand = (
                (pool['CodigoParte'] == rg['codigo']) &
                (pool['_cant'].sub(rg['cantidad']).abs() < 0.001) &
                (pool['_fob'].sub(rg['fob']).abs() <= 1)
            )
            if 'ValorFOBTotal' in pool.columns:
                mask_fob_total = (
                    pool['ValorFOBTotal'].isna() |
                    pool['ValorFOBTotal'].apply(safe_float).sub(fob_ce).abs().le(1)
                )
                mask_cand = mask_cand & mask_fob_total
            candidatos_idx = list(pool[mask_cand].index)

            if len(candidatos_idx) == 1:
                idx = candidatos_idx[0]
                df.loc[idx, 'D:CERTSM'] = nro_ce
                ya_asignadas_idx.append(idx)
                fob_acumulado = round(fob_acumulado + safe_float(df.at[idx, 'ValorTotalItem']), 2)
                pool = pool.drop(index=idx)
            else:
                renglones_pendientes.append((rg, candidatos_idx))

        # PASO 3: combinatoria solo para renglones con ambigüedad real
        for rg, candidatos_idx in renglones_pendientes:
            if not candidatos_idx:
                avisos_reparto.append(
                    f"⚠️ {nro_ce}: no se encontró línea en el Excel para "
                    f"código {rg['codigo']} (cant {rg['cantidad']}, FOB {rg['fob']})"
                )
                continue
            for idx in candidatos_idx:
                if idx in pool.index:
                    df.loc[idx, 'D:CERTSM'] = nro_ce
                    ya_asignadas_idx.append(idx)
                    fob_acumulado = round(fob_acumulado + safe_float(df.at[idx, 'ValorTotalItem']), 2)
                    pool = pool.drop(index=idx)
                    break

        fob_calc = round(fob_acumulado, 2)
        diff = abs(fob_calc - fob_ce)
        if not ya_asignadas_idx:
            estado = 'sin_match'
        else:
            estado = 'ok' if diff <= 1 else 'warn'
        resultados.append({'ce': nro_ce, 're': info['re'], 'estado': estado,
                           'fob_ce': fob_ce, 'fob_calc': fob_calc, 'n_items': len(ya_asignadas_idx)})

    if 'V:AUTOLIQCONTRIMP' not in df.columns:
        idx_certsm = df.columns.tolist().index('D:CERTSM')
        df.insert(idx_certsm + 1, 'V:AUTOLIQCONTRIMP', '')
    df['V:AUTOLIQCONTRIMP'] = df['D:CERTSM'].apply(lambda v: 'SI' if v and str(v).strip() != '' else '')

    return df, resultados, alertas_dup, avisos_reparto

# ── UI ─────────────────────────────────────────────────────────────────────────
st.title("📜 Template Unificado con CE")
st.markdown('<p class="subtitulo">INTERLOG Comercio Exterior — Asignación de Certificados Mineros al Template Unificado</p>', unsafe_allow_html=True)
st.markdown("---")

nro_ref = st.text_input("Número de referencia de la operación", placeholder="ej: 982755")

col1, col2 = st.columns([1, 1])
with col1:
    st.markdown('<p class="seccion">📊 Template Unificado</p>', unsafe_allow_html=True)
    f_unificado = st.file_uploader("Template Unificado (.xlsx)", type=["xlsx"])
with col2:
    st.markdown('<p class="seccion">📄 PDFs de CE y RE</p>', unsafe_allow_html=True)
    f_pdfs = st.file_uploader("PDFs de CE y RE (múltiples)", type=["pdf"], accept_multiple_files=True)

st.markdown("---")

if not (f_unificado and f_pdfs and nro_ref):
    st.info("📝 Ingresá el número de referencia, el Template Unificado y los PDFs de CE y RE para continuar.")

if st.button("🔍 ANALIZAR Y ASIGNAR CE", disabled=not (f_unificado and f_pdfs and nro_ref), use_container_width=True):
    with st.spinner("Procesando..."):
        df = pd.read_excel(f_unificado, dtype=str)
        df.columns = df.columns.str.strip()
        if 'CodigoParte' in df.columns:
            df['CodigoParte'] = df['CodigoParte'].apply(norm_codigo)

        ces = {}; res = {}
        for f in f_pdfs:
            fname = f.name
            if 'CE-' in fname:
                texto, nro_ce = extraer_texto_y_nro_ce(f)
                nro_ce = nro_ce or extraer_nro_ce_desde_nombre(fname)
                nro_re = extraer_re_de_ce(texto)
                if nro_ce and nro_re:
                    ces[nro_ce] = nro_re
            else:
                texto = extraer_texto_pdf(f)
                nro_re = extraer_nro_re(fname)
                fob, codigos, factura, cantidad_por_codigo, renglones = extraer_datos_re(texto)
                if nro_re:
                    res[nro_re] = {'fob': fob, 'codigos': codigos, 'factura': factura,
                                    'cantidad_por_codigo': cantidad_por_codigo,
                                    'renglones': renglones}

        ce_info = {}
        for nro_ce, nro_re in ces.items():
            if nro_re in res:
                ce_info[nro_ce] = {'re': nro_re, 'fob': res[nro_re]['fob'],
                                   'codigos': res[nro_re]['codigos'],
                                   'factura': res[nro_re]['factura'],
                                   'cantidad_por_codigo': res[nro_re]['cantidad_por_codigo'],
                                   'renglones': res[nro_re]['renglones']}

        df_resultado, resultados, alertas_dup, avisos_reparto = asignar_ce(df, ce_info)

        st.session_state.update({
            'df_resultado': df_resultado, 'resultados': resultados,
            'alertas_dup': alertas_dup, 'avisos_reparto': avisos_reparto,
            'fname_original': f_unificado.name,
            'procesado': True,
            'nro_ref': nro_ref
        })

if st.session_state.get('procesado'):
    resultados = st.session_state['resultados']
    df = st.session_state['df_resultado']
    alertas_dup = st.session_state['alertas_dup']
    fname = st.session_state['fname_original']

    if alertas_dup:
        st.markdown("### ⚠️ Alertas — CEs ambiguos")
        for a in alertas_dup:
            st.markdown(f'<span class="badge-warn">⚠️ CEs con mismos códigos y FOB: {" / ".join(a["ces"])} | FOB: {a["fob"]}</span>', unsafe_allow_html=True)

    avisos_reparto = st.session_state.get('avisos_reparto', [])
    if avisos_reparto:
        st.markdown("### ⚠️ Avisos — reparto por cantidad (códigos compartidos entre CE)")
        for a in avisos_reparto:
            st.markdown(f'<span class="badge-warn">{a}</span>', unsafe_allow_html=True)

    st.markdown("### 📊 Resultado de la asignación")
    for r in resultados:
        if r['estado'] == 'ok':
            css, badge = 'match-ok', '<span class="badge-ok">✅ OK</span>'
        elif r['estado'] == 'warn':
            css, badge = 'match-warn', '<span class="badge-warn">⚠️ Diferencia FOB</span>'
        else:
            css, badge = 'match-err', '<span class="badge-err">❌ Sin match en unificado</span>'

        st.markdown(f"""
        <div class="match-box {css}">
            {badge} <b>{r['ce']}</b><br>
            <span>RE: {r['re']}</span><br>
            <span>Ítems asignados: <b>{r['n_items']}</b> | FOB RE: <b>USD {r['fob_ce']}</b> | FOB calculado: <b>USD {r['fob_calc']}</b></span>
        </div>
        """, unsafe_allow_html=True)

    sin_ce = df[(df['D:CERTSM'] == '') & (df['Observaciones'].isna() | (df['Observaciones'].str.strip() == ''))]
    total_sin_cm = (df['Observaciones'].str.strip() == 'SIN CM').sum() if 'Observaciones' in df.columns else 0
    total_asignados = (df['D:CERTSM'] != '').sum()

    if not sin_ce.empty:
        st.markdown(f'<br><span class="badge-warn">⚠️ {len(sin_ce)} ítems sin CE asignado</span>', unsafe_allow_html=True)
        st.dataframe(sin_ce[['NumeroDeFactura','CodigoParte','ValorTotalItem','ITEM']].reset_index(drop=True))

    st.markdown(f"**Total: {total_asignados} ítems con CE | {len(sin_ce)} sin CE | {total_sin_cm} SIN CM**")
    st.markdown("---")

    st.download_button(
        label="📥 DESCARGAR TEMPLATE UNIFICADO CON CE",
        data=exportar_excel(df),
        file_name=f'TEMPLATE_UNIFICADO_{st.session_state.get("nro_ref", "")}_CON_CE.xlsx',
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True
    )
